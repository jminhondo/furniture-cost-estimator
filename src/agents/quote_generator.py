"""
Quote Generator — Capa 4.

Produces three outputs from Layer-2 (BOM) + Layer-3 (costs) JSON:

1. PDF via reportlab
   - Two modes: ``resumido`` (one item per module) and ``detallado``
     (one item per cost category / herraje line)
   - Header with company data, metadata grid, items table, totals,
     three price scenarios, footer with payment conditions
   - Quote number: PRES-YYYYMM-XXXX (sequential counter per output dir)

2. Excel via openpyxl
   - Sheet "BOM":         all cut pieces with dimensions
   - Sheet "Herrajes":    hardware lines with prices
   - Sheet "Plan de corte": board layout positions
   - Sheet "Costos":      full cost breakdown + final price
   - Bold headers, alternating row shading

3. CRM push via httpx
   - Endpoint from env var CRM_ENDPOINT (or config crm.endpoint)
   - Payload: proyecto_nombre, cliente_id, monto, estado=draft, numero
   - Non-fatal: if CRM fails, generation continues; error is logged

Input (``quote_data`` dict)
---------------------------
{
    "proyecto_id": str,
    "proyecto_nombre": str,
    "cliente": {
        "nombre": str,
        "cliente_id": str,
        "direccion_obra": str,
        "email": str,
        "telefono": str,
    },
    "layer2": dict,   # BOMBuilder.build() output
    "layer3": dict,   # CostEngine.estimate() output
}

Output
------
{
    "quote_number": "PRES-202603-0001",
    "files": {
        "pdf_resumido": "/abs/path/PRES-202603-0001_resumido.pdf",
        "pdf_detallado": "/abs/path/PRES-202603-0001_detallado.pdf",
        "excel": "/abs/path/PRES-202603-0001.xlsx",
    },
    "crm": {"success": bool, "status_code": int|None, "error": str|None},
}
"""

from __future__ import annotations

import logging
import os
from datetime import date
from pathlib import Path
from typing import Literal

import httpx

# reportlab ----------------------------------------------------------------
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import (
    HRFlowable,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)

# openpyxl -----------------------------------------------------------------
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Brand palette
# ---------------------------------------------------------------------------
_NAVY = colors.HexColor("#1B3A5C")
_LIGHT_BLUE = colors.HexColor("#E8F0FE")
_MID_GRAY = colors.HexColor("#D0D0D0")
_WHITE = colors.white

# Excel fills
_XL_HEADER_FILL = PatternFill(fgColor="1B3A5C", fill_type="solid")
_XL_ALT_FILL = PatternFill(fgColor="E8F0FE", fill_type="solid")
_XL_HEADER_FONT = Font(bold=True, color="FFFFFF")
_XL_BOLD = Font(bold=True)
_XL_BORDER_SIDE = Side(style="thin", color="D0D0D0")
_XL_BORDER = Border(
    left=_XL_BORDER_SIDE, right=_XL_BORDER_SIDE,
    top=_XL_BORDER_SIDE, bottom=_XL_BORDER_SIDE,
)

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _usd(value: float) -> str:
    return f"$ {value:,.2f}"


def _pct(value: float) -> str:
    return f"{value * 100:.0f}%"


# ---------------------------------------------------------------------------
# QuoteGenerator
# ---------------------------------------------------------------------------

class QuoteGenerator:
    """
    Generate PDF, Excel and CRM push for a furniture project quote.

    Parameters
    ----------
    config:     Pre-loaded config dict (from defaults.yml).
    output_dir: Directory where all generated files are written.
                Created automatically if it doesn't exist.
    """

    def __init__(self, config: dict, output_dir: str | Path = ".") -> None:
        self._cfg = config
        self._out = Path(output_dir)
        self._out.mkdir(parents=True, exist_ok=True)

        empresa_cfg = config.get("empresa", {})
        self._empresa = {
            "nombre": empresa_cfg.get("nombre") or os.getenv("EMPRESA_NOMBRE", "Carpintería"),
            "direccion": empresa_cfg.get("direccion", ""),
            "telefono": empresa_cfg.get("telefono", ""),
            "email": empresa_cfg.get("email", ""),
            "rut": empresa_cfg.get("rut", ""),
            "condiciones_pago": empresa_cfg.get("condiciones_pago", "A convenir"),
            "garantia": empresa_cfg.get("garantia", "12 meses en estructura"),
        }

        precio_cfg = config.get("precio", {})
        self._iva: float = float(precio_cfg.get("iva_pct", config.get("iva_pct", 0.21)))
        self._factor_cons: float = float(precio_cfg.get("escenario_conservador_factor", 0.90))
        self._factor_agr: float = float(precio_cfg.get("escenario_agresivo_factor", 1.08))

        crm_cfg = config.get("crm", {})
        self._crm_endpoint: str = os.getenv("CRM_ENDPOINT", crm_cfg.get("endpoint", ""))
        self._crm_timeout: float = float(crm_cfg.get("timeout_s", 10))

    # ------------------------------------------------------------------
    # Public API
    # ------------------------------------------------------------------

    async def generate(self, quote_data: dict) -> dict:
        """
        Generate all three outputs and return paths + CRM result.

        This method is async because CRM push uses httpx async client.
        PDF and Excel generation are synchronous (CPU/IO bound).
        """
        quote_num = self._next_quote_number()
        layer2: dict = quote_data.get("layer2", {})
        layer3: dict = quote_data.get("layer3", {})

        # PDF — both modes
        pdf_res = self._generate_pdf(quote_data, quote_num, layer2, layer3, "resumido")
        pdf_det = self._generate_pdf(quote_data, quote_num, layer2, layer3, "detallado")

        # Excel
        xlsx = self._generate_excel(quote_data, quote_num, layer2, layer3)

        # CRM
        crm = await self._push_crm(quote_data, quote_num, layer3)

        return {
            "quote_number": quote_num,
            "files": {
                "pdf_resumido": str(pdf_res),
                "pdf_detallado": str(pdf_det),
                "excel": str(xlsx),
            },
            "crm": crm,
        }

    # ------------------------------------------------------------------
    # Quote number
    # ------------------------------------------------------------------

    def _next_quote_number(self) -> str:
        """Return a sequential PRES-YYYYMM-XXXX number, persisted per output_dir."""
        counter_file = self._out / ".quote_counter"
        n = 1
        if counter_file.exists():
            try:
                n = int(counter_file.read_text().strip()) + 1
            except ValueError:
                n = 1
        counter_file.write_text(str(n))
        ym = date.today().strftime("%Y%m")
        return f"PRES-{ym}-{n:04d}"

    # ------------------------------------------------------------------
    # PDF generation
    # ------------------------------------------------------------------

    def _generate_pdf(
        self,
        quote_data: dict,
        quote_num: str,
        layer2: dict,
        layer3: dict,
        mode: Literal["resumido", "detallado"],
    ) -> Path:
        """Build a PDF for the given mode and write it to output_dir."""
        path = self._out / f"{quote_num}_{mode}.pdf"
        cliente = quote_data.get("cliente", {})
        proyecto_nombre = quote_data.get("proyecto_nombre", quote_data.get("proyecto_id", "—"))

        doc = SimpleDocTemplate(
            str(path),
            pagesize=A4,
            leftMargin=2 * cm, rightMargin=2 * cm,
            topMargin=1.5 * cm, bottomMargin=1.5 * cm,
        )

        styles = getSampleStyleSheet()
        s_normal = styles["Normal"]
        s_h1 = ParagraphStyle("h1", fontSize=18, textColor=_NAVY, spaceAfter=2,
                               fontName="Helvetica-Bold")
        s_h2 = ParagraphStyle("h2", fontSize=11, textColor=_NAVY, spaceAfter=4,
                               fontName="Helvetica-Bold")
        s_small = ParagraphStyle("small", fontSize=8, textColor=colors.gray,
                                  fontName="Helvetica")
        s_footer = ParagraphStyle("footer", fontSize=7.5, textColor=colors.gray,
                                   fontName="Helvetica", leading=11)
        s_cell = ParagraphStyle("cell", fontSize=9, fontName="Helvetica",
                                 wordWrap="CJK")

        story = []

        # ── Header ──────────────────────────────────────────────────────
        header_data = [
            [
                Paragraph(self._empresa["nombre"], s_h1),
                Paragraph(
                    f"<b>PRESUPUESTO</b>",
                    ParagraphStyle("pr", fontSize=14, textColor=_NAVY,
                                   fontName="Helvetica-Bold", alignment=2),
                ),
            ],
            [
                Paragraph(
                    f"{self._empresa['direccion']}<br/>"
                    f"Tel: {self._empresa['telefono']} · {self._empresa['email']}<br/>"
                    f"RUT: {self._empresa['rut']}",
                    s_small,
                ),
                Paragraph(
                    f"<b>Nro:</b> {quote_num}<br/>"
                    f"<b>Fecha:</b> {date.today().strftime('%d/%m/%Y')}",
                    ParagraphStyle("pr2", fontSize=9, fontName="Helvetica",
                                   alignment=2),
                ),
            ],
        ]
        header_table = Table(header_data, colWidths=["60%", "40%"])
        header_table.setStyle(TableStyle([("VALIGN", (0, 0), (-1, -1), "TOP")]))
        story.append(header_table)
        story.append(Spacer(1, 0.3 * cm))
        story.append(HRFlowable(width="100%", thickness=2, color=_NAVY))
        story.append(Spacer(1, 0.4 * cm))

        # ── Metadata grid ────────────────────────────────────────────────
        story.append(Paragraph("DATOS DEL PROYECTO", s_h2))
        meta_data = [
            ["Cliente:", cliente.get("nombre", "—"),
             "Proyecto:", proyecto_nombre],
            ["E-mail:", cliente.get("email", "—"),
             "Teléfono:", cliente.get("telefono", "—")],
            ["Dirección de obra:", Paragraph(cliente.get("direccion_obra", "—"), s_cell),
             "", ""],
        ]
        meta_table = Table(meta_data, colWidths=["18%", "32%", "18%", "32%"])
        meta_table.setStyle(TableStyle([
            ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
            ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
            ("FONTNAME", (2, 0), (2, -1), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("ROWBACKGROUNDS", (0, 0), (-1, -1), [_LIGHT_BLUE, _WHITE]),
            ("BOX", (0, 0), (-1, -1), 0.5, _MID_GRAY),
            ("INNERGRID", (0, 0), (-1, -1), 0.25, _MID_GRAY),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ]))
        story.append(meta_table)
        story.append(Spacer(1, 0.5 * cm))

        # ── Items table ──────────────────────────────────────────────────
        mode_label = "RESUMEN DE ITEMS" if mode == "resumido" else "DETALLE DE COSTOS"
        story.append(Paragraph(mode_label, s_h2))

        if mode == "resumido":
            rows = self._items_resumido(layer2, layer3)
        else:
            rows = self._items_detallado(layer2, layer3)

        col_w = ["52%", "12%", "18%", "18%"]
        items_header = [["Descripción", "Cant.", "Precio unit.", "Total"]]
        items_data = items_header + [
            [Paragraph(r[0], s_cell), r[1], r[2], r[3]] for r in rows
        ]
        items_table = Table(items_data, colWidths=col_w)
        items_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), _NAVY),
            ("TEXTCOLOR", (0, 0), (-1, 0), _WHITE),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [_WHITE, _LIGHT_BLUE]),
            ("ALIGN", (1, 0), (-1, -1), "RIGHT"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("BOX", (0, 0), (-1, -1), 0.5, _MID_GRAY),
            ("INNERGRID", (0, 0), (-1, -1), 0.25, _MID_GRAY),
            ("TOPPADDING", (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ]))
        story.append(items_table)
        story.append(Spacer(1, 0.4 * cm))

        # ── Totals ───────────────────────────────────────────────────────
        precio_base: float = float(layer3.get("precio_sugerido_usd", 0))
        costo_total: float = float(layer3.get("costo_total_usd", 0))
        subtotal = costo_total
        iva_amt = subtotal * self._iva
        total_con_iva = subtotal + iva_amt

        totals_data = [
            ["", "Subtotal:", _usd(subtotal)],
            ["", f"IVA ({_pct(self._iva)}):", _usd(iva_amt)],
            ["", "TOTAL:", _usd(total_con_iva)],
        ]
        totals_table = Table(totals_data, colWidths=["52%", "28%", "20%"])
        totals_table.setStyle(TableStyle([
            ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
            ("FONTNAME", (1, 0), (-1, -1), "Helvetica-Bold"),
            ("FONTNAME", (1, 2), (-1, 2), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 2), 10),
            ("FONTSIZE", (0, 2), (-1, 2), 12),
            ("TEXTCOLOR", (0, 2), (-1, 2), _NAVY),
            ("ALIGN", (1, 0), (-1, -1), "RIGHT"),
            ("LINEABOVE", (1, 2), (-1, 2), 1.5, _NAVY),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ]))
        story.append(totals_table)
        story.append(Spacer(1, 0.5 * cm))

        # ── Price scenarios ───────────────────────────────────────────────
        story.append(Paragraph("ESCENARIOS DE PRECIO", s_h2))
        p_cons = precio_base * self._factor_cons
        p_agr = precio_base * self._factor_agr
        scenarios_data = [
            ["Escenario", "Variación", "Precio final"],
            ["Conservador", f"-{_pct(1 - self._factor_cons)}", _usd(p_cons)],
            ["Base (recomendado)", "—", _usd(precio_base)],
            ["Agresivo", f"+{_pct(self._factor_agr - 1)}", _usd(p_agr)],
        ]
        sc_table = Table(scenarios_data, colWidths=["40%", "25%", "35%"])
        sc_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), _NAVY),
            ("TEXTCOLOR", (0, 0), (-1, 0), _WHITE),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("BACKGROUND", (0, 2), (-1, 2), _LIGHT_BLUE),
            ("FONTNAME", (0, 2), (-1, 2), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("ALIGN", (1, 0), (-1, -1), "CENTER"),
            ("BOX", (0, 0), (-1, -1), 0.5, _MID_GRAY),
            ("INNERGRID", (0, 0), (-1, -1), 0.25, _MID_GRAY),
            ("TOPPADDING", (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ]))
        story.append(sc_table)
        story.append(Spacer(1, 0.6 * cm))

        # ── Footer ────────────────────────────────────────────────────────
        story.append(HRFlowable(width="100%", thickness=1, color=_MID_GRAY))
        story.append(Spacer(1, 0.2 * cm))
        story.append(Paragraph(
            f"<b>Condiciones de pago:</b> {self._empresa['condiciones_pago']}  |  "
            f"<b>Garantía:</b> {self._empresa['garantia']}",
            s_footer,
        ))
        story.append(Paragraph(
            "Presupuesto válido por 30 días desde la fecha de emisión. "
            "Precios en USD. No incluye IVA salvo indicación.",
            s_footer,
        ))

        doc.build(story)
        logger.info("PDF (%s) written to %s", mode, path)
        return path

    # ------------------------------------------------------------------
    # Items table data (exposed for testing)
    # ------------------------------------------------------------------

    def _items_resumido(self, layer2: dict, layer3: dict) -> list[tuple[str, str, str, str]]:
        """
        One row per module (material group) + one row per cost agent.

        Returns list of (descripcion, cantidad, precio_unit, total) tuples.
        """
        rows: list[tuple[str, str, str, str]] = []

        # Tableros — one row per material group
        for tablero in layer2.get("tableros", []):
            mat_id = tablero.get("material_id", "Material")
            n = tablero.get("tableros_necesarios", 0)
            costo = float(tablero.get("costo_tableros_usd", 0))
            rows.append((
                f"Tablero {mat_id.replace('_', ' ').title()} ({n} ud.)",
                "1",
                _usd(costo),
                _usd(costo),
            ))

        # Herrajes — single consolidated row
        total_herr = float(layer3.get("costo_materiales_usd", 0)) - sum(
            float(t.get("costo_tableros_usd", 0)) for t in layer2.get("tableros", [])
        )
        if total_herr > 0:
            rows.append(("Herrajes y accesorios", "1", _usd(total_herr), _usd(total_herr)))

        # Cost agents
        labor = float(layer3.get("costo_mano_obra_usd", 0))
        install = float(layer3.get("costo_instalacion_usd", 0))
        transport = float(layer3.get("costo_transporte_usd", 0))
        overhead = float(layer3.get("costo_indirecto_usd", 0))

        if labor > 0:
            rows.append(("Mano de obra — fabricación", "1", _usd(labor), _usd(labor)))
        if install > 0:
            rows.append(("Instalación en obra", "1", _usd(install), _usd(install)))
        if transport > 0:
            rows.append(("Transporte", "1", _usd(transport), _usd(transport)))
        if overhead > 0:
            rows.append(("Gastos indirectos (overhead)", "1", _usd(overhead), _usd(overhead)))

        return rows

    def _items_detallado(self, layer2: dict, layer3: dict) -> list[tuple[str, str, str, str]]:
        """
        Expanded rows: one per herraje line, then cost categories.

        Returns list of (descripcion, cantidad, precio_unit, total) tuples.
        """
        rows: list[tuple[str, str, str, str]] = []

        # Tableros — one row per material group (same as resumido)
        for tablero in layer2.get("tableros", []):
            mat_id = tablero.get("material_id", "Material")
            n = tablero.get("tableros_necesarios", 0)
            piezas = len(tablero.get("plan_de_corte", []))
            costo = float(tablero.get("costo_tableros_usd", 0))
            desperdicio = tablero.get("desperdicio_pct", 0)
            rows.append((
                f"Tablero {mat_id.replace('_', ' ').title()} · "
                f"{n} hoja(s), {piezas} piezas, desperdc. {desperdicio*100:.1f}%",
                str(n),
                _usd(costo / n if n else 0),
                _usd(costo),
            ))

        # Herrajes — one row per line item
        for linea in layer2.get("herrajes", {}).get("lineas", []):
            nombre = linea.get("nombre", linea.get("herraje_id", "Herraje"))
            cant = linea.get("cantidad", 0)
            unidad = linea.get("unidad", "u")
            p_unit = float(linea.get("precio_unitario_usd", 0))
            p_total = float(linea.get("precio_total_usd", 0))
            cant_str = f"{cant} {unidad}" if unidad == "ml" else str(cant)
            rows.append((nombre, cant_str, _usd(p_unit), _usd(p_total)))

        # Cost categories from layer3
        cats = [
            ("costo_mano_obra_usd", "Mano de obra — fabricación"),
            ("costo_instalacion_usd", "Instalación en obra"),
            ("costo_transporte_usd", "Transporte de materiales y producto"),
            ("costo_indirecto_usd", "Gastos indirectos — overhead"),
        ]
        for key, label in cats:
            val = float(layer3.get(key, 0))
            if val > 0:
                rows.append((label, "1", _usd(val), _usd(val)))

        return rows

    # ------------------------------------------------------------------
    # Excel generation
    # ------------------------------------------------------------------

    def _generate_excel(
        self,
        quote_data: dict,
        quote_num: str,
        layer2: dict,
        layer3: dict,
    ) -> Path:
        """Write a 4-sheet workbook and return the file path."""
        path = self._out / f"{quote_num}.xlsx"
        wb = Workbook()
        wb.remove(wb.active)  # remove default sheet

        self._xl_sheet_bom(wb, layer2)
        self._xl_sheet_herrajes(wb, layer2)
        self._xl_sheet_plan_corte(wb, layer2)
        self._xl_sheet_costos(wb, layer3, quote_data)

        wb.save(str(path))
        logger.info("Excel written to %s", path)
        return path

    def _xl_write_header(self, ws, columns: list[str]) -> None:
        """Write a styled header row."""
        for col_idx, title in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=title)
            cell.font = _XL_HEADER_FONT
            cell.fill = _XL_HEADER_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = _XL_BORDER

    def _xl_write_row(self, ws, row_idx: int, values: list, alternate: bool) -> None:
        """Write a data row with optional alternate shading."""
        fill = _XL_ALT_FILL if alternate else None
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = _XL_BORDER
            if fill:
                cell.fill = fill
            if isinstance(val, float):
                cell.number_format = '#,##0.00'

    def _xl_sheet_bom(self, wb: Workbook, layer2: dict) -> None:
        ws = wb.create_sheet("BOM")
        cols = ["Pieza ID", "Nombre", "Material", "Tablero #", "Largo (mm)", "Ancho (mm)", "Rotada"]
        self._xl_write_header(ws, cols)
        ws.column_dimensions["A"].width = 14
        ws.column_dimensions["B"].width = 30
        ws.column_dimensions["C"].width = 18

        row = 2
        for tablero in layer2.get("tableros", []):
            mat_id = tablero.get("material_id", "")
            for pieza in tablero.get("plan_de_corte", []):
                self._xl_write_row(ws, row, [
                    pieza.get("pieza_id", ""),
                    pieza.get("nombre", ""),
                    mat_id,
                    pieza.get("tablero_num", ""),
                    pieza.get("largo_mm", ""),
                    pieza.get("ancho_mm", ""),
                    "Sí" if pieza.get("rotada") else "No",
                ], alternate=(row % 2 == 0))
                row += 1

    def _xl_sheet_herrajes(self, wb: Workbook, layer2: dict) -> None:
        ws = wb.create_sheet("Herrajes")
        cols = ["Herraje ID", "Nombre", "Cantidad", "Unidad",
                "Precio unit. (USD)", "Precio total (USD)"]
        self._xl_write_header(ws, cols)
        ws.column_dimensions["B"].width = 35
        ws.column_dimensions["E"].width = 18
        ws.column_dimensions["F"].width = 18

        row = 2
        for linea in layer2.get("herrajes", {}).get("lineas", []):
            self._xl_write_row(ws, row, [
                linea.get("herraje_id", ""),
                linea.get("nombre", ""),
                linea.get("cantidad", ""),
                linea.get("unidad", ""),
                float(linea.get("precio_unitario_usd", 0)),
                float(linea.get("precio_total_usd", 0)),
            ], alternate=(row % 2 == 0))
            row += 1

        # Total row
        total = float(layer2.get("herrajes", {}).get("total_herrajes_usd", 0))
        total_cell = ws.cell(row=row, column=6, value=total)
        total_cell.font = _XL_BOLD
        total_cell.number_format = '#,##0.00'
        label_cell = ws.cell(row=row, column=5, value="TOTAL:")
        label_cell.font = _XL_BOLD
        label_cell.alignment = Alignment(horizontal="right")

    def _xl_sheet_plan_corte(self, wb: Workbook, layer2: dict) -> None:
        ws = wb.create_sheet("Plan de corte")
        cols = ["Material", "Tablero #", "Pieza", "X (mm)", "Y (mm)",
                "Largo (mm)", "Ancho (mm)", "Rotada"]
        self._xl_write_header(ws, cols)
        ws.column_dimensions["A"].width = 18
        ws.column_dimensions["C"].width = 28

        row = 2
        for tablero in layer2.get("tableros", []):
            mat_id = tablero.get("material_id", "")
            for pieza in tablero.get("plan_de_corte", []):
                self._xl_write_row(ws, row, [
                    mat_id,
                    pieza.get("tablero_num", ""),
                    pieza.get("nombre", ""),
                    pieza.get("x_mm", ""),
                    pieza.get("y_mm", ""),
                    pieza.get("largo_mm", ""),
                    pieza.get("ancho_mm", ""),
                    "Sí" if pieza.get("rotada") else "No",
                ], alternate=(row % 2 == 0))
                row += 1

    def _xl_sheet_costos(self, wb: Workbook, layer3: dict, quote_data: dict) -> None:
        ws = wb.create_sheet("Costos")
        ws.column_dimensions["A"].width = 38
        ws.column_dimensions["B"].width = 20

        cost_rows = [
            ("Materiales (tableros + herrajes)", layer3.get("costo_materiales_usd", 0)),
            ("Mano de obra — fabricación", layer3.get("costo_mano_obra_usd", 0)),
            ("Instalación en obra", layer3.get("costo_instalacion_usd", 0)),
            ("Transporte", layer3.get("costo_transporte_usd", 0)),
            ("Gastos indirectos (overhead)", layer3.get("costo_indirecto_usd", 0)),
            ("─" * 30, None),
            ("Costo directo", layer3.get("costo_directo_usd", 0)),
            ("Costo total", layer3.get("costo_total_usd", 0)),
            ("Precio sugerido (con margen)", layer3.get("precio_sugerido_usd", 0)),
        ]

        # Header
        ws.cell(row=1, column=1, value="Concepto").font = _XL_BOLD
        ws.cell(row=1, column=2, value="Monto (USD)").font = _XL_BOLD

        for i, (label, value) in enumerate(cost_rows, 2):
            label_cell = ws.cell(row=i, column=1, value=label)
            if value is not None:
                val_cell = ws.cell(row=i, column=2, value=float(value))
                val_cell.number_format = '#,##0.00'
            alternate = i % 2 == 0
            if alternate:
                label_cell.fill = _XL_ALT_FILL
                if value is not None:
                    val_cell.fill = _XL_ALT_FILL

        # Highlight precio sugerido
        last_row = 1 + len(cost_rows)
        ws.cell(row=last_row, column=1).font = Font(bold=True, color="1B3A5C")
        ws.cell(row=last_row, column=2).font = Font(bold=True, color="1B3A5C")

        # Confidence note
        confianza = layer3.get("confianza_global", 0)
        note_row = last_row + 2
        note_cell = ws.cell(row=note_row, column=1,
                            value=f"Confianza del modelo: {confianza*100:.1f}%")
        note_cell.font = Font(italic=True, color="808080")

    # ------------------------------------------------------------------
    # CRM push
    # ------------------------------------------------------------------

    async def _push_crm(self, quote_data: dict, quote_num: str, layer3: dict) -> dict:
        """
        POST quote summary to the CRM endpoint.

        Returns a result dict regardless of success/failure — CRM errors
        are non-fatal (logged only).
        """
        if not self._crm_endpoint:
            logger.debug("CRM_ENDPOINT not configured — skipping CRM push")
            return {"success": False, "status_code": None, "error": "no endpoint configured"}

        payload = {
            "numero_presupuesto": quote_num,
            "nombre": quote_data.get("proyecto_nombre", quote_data.get("proyecto_id", "")),
            "cliente_id": quote_data.get("cliente", {}).get("cliente_id", ""),
            "monto": float(layer3.get("precio_sugerido_usd", 0)),
            "estado": "draft",
        }

        try:
            async with httpx.AsyncClient(timeout=self._crm_timeout) as client:
                response = await client.post(self._crm_endpoint, json=payload)
                response.raise_for_status()
                logger.info("CRM push OK — quote %s, status %d", quote_num, response.status_code)
                return {"success": True, "status_code": response.status_code, "error": None}
        except httpx.HTTPStatusError as exc:
            logger.warning("CRM HTTP error %d for quote %s", exc.response.status_code, quote_num)
            return {"success": False, "status_code": exc.response.status_code,
                    "error": str(exc)}
        except httpx.RequestError as exc:
            logger.warning("CRM connection error for quote %s: %s", quote_num, exc)
            return {"success": False, "status_code": None, "error": str(exc)}
