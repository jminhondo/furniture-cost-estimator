"""
Unit tests for QuoteGenerator.

Tests
-----
1. test_pdf_generates_file              — both PDF files are created on disk
2. test_pdf_resumido_has_module_items   — resumido rows contain material names
3. test_pdf_detallado_has_cost_categories — detallado rows contain cost labels
4. test_excel_has_four_sheets           — workbook has the expected 4 sheets
5. test_crm_push_handles_connection_error — ConnectError → success=False, no raise
"""

from __future__ import annotations

import pytest
from openpyxl import load_workbook
from unittest.mock import AsyncMock, MagicMock, patch

import httpx

from src.agents.quote_generator import QuoteGenerator

# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------

_TEST_CONFIG: dict = {
    "empresa": {
        "nombre": "Test Carpintería S.A.",
        "direccion": "Calle Falsa 123",
        "telefono": "+54 11 0000-0000",
        "email": "test@example.com",
        "rut": "00-000-0",
        "condiciones_pago": "50% anticipo",
        "garantia": "12 meses en estructura",
    },
    "precio": {
        "margen_default": 0.32,
        "iva_pct": 0.21,
        "escenario_conservador_factor": 0.90,
        "escenario_agresivo_factor": 1.08,
    },
    "crm": {"endpoint": "", "timeout_s": 5},
}

_LAYER2: dict = {
    "proyecto_id": "test-001",
    "tableros": [
        {
            "material_id": "mdf_18",
            "tableros_necesarios": 3,
            "costo_tableros_usd": 207.56,
            "desperdicio_pct": 0.12,
            "plan_de_corte": [
                {
                    "pieza_id": "p1", "nombre": "Lateral Izquierdo",
                    "tablero_num": 0, "x_mm": 0, "y_mm": 0,
                    "largo_mm": 700.0, "ancho_mm": 400.0, "rotada": False,
                },
                {
                    "pieza_id": "p2", "nombre": "Puerta Principal",
                    "tablero_num": 1, "x_mm": 0, "y_mm": 750,
                    "largo_mm": 700.0, "ancho_mm": 500.0, "rotada": False,
                },
                {
                    "pieza_id": "p3", "nombre": "Techo",
                    "tablero_num": 2, "x_mm": 0, "y_mm": 0,
                    "largo_mm": 900.0, "ancho_mm": 580.0, "rotada": False,
                },
            ],
        },
    ],
    "herrajes": {
        "lineas": [
            {
                "herraje_id": "bisagra_blum_110",
                "nombre": "Bisagra Blum 110° soft-close",
                "cantidad": 4,
                "precio_unitario_usd": 4.50,
                "precio_total_usd": 18.00,
                "unidad": "u",
            },
            {
                "herraje_id": "canto_pvc",
                "nombre": "Canto PVC 0.4mm",
                "cantidad": 12.5,
                "precio_unitario_usd": 0.85,
                "precio_total_usd": 10.63,
                "unidad": "ml",
            },
        ],
        "total_herrajes_usd": 28.63,
    },
    "resumen": {"total_materiales_usd": 207.56, "piezas_procesadas": 8},
}

_LAYER3: dict = {
    "proyecto_id": "test-001",
    "costo_materiales_usd": 236.19,
    "costo_mano_obra_usd": 320.00,
    "costo_instalacion_usd": 180.00,
    "costo_transporte_usd": 85.00,
    "costo_indirecto_usd": 146.00,
    "costo_directo_usd": 821.19,
    "costo_total_usd": 967.19,
    "precio_sugerido_usd": 1537.26,
    "confianza_global": 0.74,
    "fuentes": {
        "materiales": "bom",
        "mano_obra": "rules",
        "instalacion": "rules",
        "transporte": "rules",
        "indirecto": "rules",
    },
    "intervalos": {
        "labor_low": 256.0, "labor_high": 400.0,
        "install_low": 144.0, "install_high": 234.0,
        "transport_low": 72.0, "transport_high": 102.0,
        "overhead_low": 131.0, "overhead_high": 168.0,
    },
}

_QUOTE_DATA: dict = {
    "proyecto_id": "test-001",
    "proyecto_nombre": "Cocina Principal",
    "cliente": {
        "nombre": "Juan Pérez",
        "cliente_id": "CL-001",
        "direccion_obra": "Av. Corrientes 1234, CABA",
        "email": "juan@example.com",
        "telefono": "+54 11 1234-5678",
    },
    "layer2": _LAYER2,
    "layer3": _LAYER3,
}


@pytest.fixture
def gen(tmp_path):
    return QuoteGenerator(_TEST_CONFIG, output_dir=tmp_path)


# ---------------------------------------------------------------------------
# Test 1 — PDF files are created
# ---------------------------------------------------------------------------

@pytest.mark.asyncio
async def test_pdf_generates_file(gen, tmp_path):
    result = await gen.generate(_QUOTE_DATA)

    pdf_res = tmp_path / result["files"]["pdf_resumido"].split("/")[-1]
    pdf_det = tmp_path / result["files"]["pdf_detallado"].split("/")[-1]

    assert pdf_res.exists(), "Resumido PDF not found"
    assert pdf_det.exists(), "Detallado PDF not found"
    # Valid PDF starts with %PDF
    assert pdf_res.read_bytes()[:4] == b"%PDF"
    assert pdf_det.read_bytes()[:4] == b"%PDF"


# ---------------------------------------------------------------------------
# Test 2 — Resumido rows contain material names
# ---------------------------------------------------------------------------

def test_pdf_resumido_has_module_items(gen):
    rows = gen._items_resumido(_LAYER2, _LAYER3)

    descriptions = [r[0] for r in rows]
    # Material group should appear
    assert any("mdf" in d.lower() or "Mdf" in d for d in descriptions), \
        f"Expected 'mdf_18' material in resumido rows: {descriptions}"
    # Labor should appear
    assert any("mano de obra" in d.lower() for d in descriptions), \
        f"Expected labor row: {descriptions}"
    # Each row has 4 columns
    for row in rows:
        assert len(row) == 4


# ---------------------------------------------------------------------------
# Test 3 — Detallado rows contain cost categories and herraje line items
# ---------------------------------------------------------------------------

def test_pdf_detallado_has_cost_categories(gen):
    rows = gen._items_detallado(_LAYER2, _LAYER3)

    descriptions = [r[0] for r in rows]

    # Individual herraje lines should appear
    assert any("bisagra" in d.lower() for d in descriptions), \
        f"Expected bisagra line in detallado rows: {descriptions}"
    assert any("canto" in d.lower() for d in descriptions), \
        f"Expected canto PVC line: {descriptions}"

    # Cost categories
    assert any("mano de obra" in d.lower() for d in descriptions)
    assert any("instalación" in d.lower() or "instalacion" in d.lower() for d in descriptions)
    assert any("transporte" in d.lower() for d in descriptions)
    assert any("overhead" in d.lower() or "indirecto" in d.lower() for d in descriptions)


# ---------------------------------------------------------------------------
# Test 4 — Excel has exactly 4 sheets with correct names
# ---------------------------------------------------------------------------

@pytest.mark.asyncio
async def test_excel_has_four_sheets(gen, tmp_path):
    result = await gen.generate(_QUOTE_DATA)

    xlsx_path = result["files"]["excel"]
    wb = load_workbook(xlsx_path)

    expected_sheets = {"BOM", "Herrajes", "Plan de corte", "Costos"}
    assert set(wb.sheetnames) == expected_sheets, \
        f"Sheet names mismatch: {wb.sheetnames}"

    # BOM sheet has data rows
    bom_ws = wb["BOM"]
    assert bom_ws.max_row > 1, "BOM sheet should have at least one data row"

    # Herrajes sheet has data rows
    herr_ws = wb["Herrajes"]
    assert herr_ws.max_row > 1

    # Costos sheet has precio_sugerido
    costos_ws = wb["Costos"]
    values = [costos_ws.cell(row=r, column=1).value for r in range(1, costos_ws.max_row + 1)]
    assert any(v and "sugerido" in str(v).lower() for v in values), \
        f"Costos sheet should mention 'sugerido': {values}"


# ---------------------------------------------------------------------------
# Test 5 — CRM push handles connection errors gracefully
# ---------------------------------------------------------------------------

@pytest.mark.asyncio
async def test_crm_push_handles_connection_error(tmp_path):
    config = {**_TEST_CONFIG, "crm": {"endpoint": "http://crm.example.com/api", "timeout_s": 5}}
    gen = QuoteGenerator(config, output_dir=tmp_path)

    with patch("httpx.AsyncClient.post", new_callable=AsyncMock) as mock_post:
        mock_post.side_effect = httpx.ConnectError("Connection refused")
        crm = await gen._push_crm(_QUOTE_DATA, "PRES-202603-0001", _LAYER3)

    assert crm["success"] is False
    assert crm["status_code"] is None
    assert crm["error"] is not None


# ---------------------------------------------------------------------------
# Bonus: quote number format
# ---------------------------------------------------------------------------

def test_quote_number_format(gen):
    num = gen._next_quote_number()
    parts = num.split("-")
    assert parts[0] == "PRES"
    assert len(parts[1]) == 6  # YYYYMM
    assert len(parts[2]) == 4  # XXXX zero-padded
    assert parts[2].isdigit()


def test_quote_number_increments(gen):
    n1 = gen._next_quote_number()
    n2 = gen._next_quote_number()
    seq1 = int(n1.split("-")[2])
    seq2 = int(n2.split("-")[2])
    assert seq2 == seq1 + 1
